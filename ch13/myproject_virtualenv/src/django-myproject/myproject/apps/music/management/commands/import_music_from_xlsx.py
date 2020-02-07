from django.core.management.base import BaseCommand


class Command(BaseCommand):
    help = (
        "Imports music from a local XLSX file. "
        "Expects columns: Artist, Title, URL"
    )
    SILENT, NORMAL, VERBOSE, VERY_VERBOSE = 0, 1, 2, 3

    def add_arguments(self, parser):
        # Positional arguments
        parser.add_argument("file_path",
                            nargs=1,
                            type=str)

    def handle(self, *args, **options):
        self.verbosity = options.get("verbosity", self.NORMAL)
        self.file_path = options["file_path"][0]
        self.prepare()
        self.main()
        self.finalize()

    def prepare(self):
        self.imported_counter = 0
        self.skipped_counter = 0

    def main(self):
        from openpyxl import load_workbook
        from ...forms import SongForm

        wb = load_workbook(filename=self.file_path)
        ws = wb.worksheets[0]

        if self.verbosity >= self.NORMAL:
            self.stdout.write("=== Importing music ===")

        columns = ["artist", "title", "url"]
        rows = ws.iter_rows(min_row=2)  # skip the column captions
        for index, row in enumerate(rows, start=1):
            row_values = [cell.value for cell in row]
            row_dict = dict(zip(columns, row_values))
            form = SongForm(data=row_dict)
            if form.is_valid():
                song = form.save()
                if self.verbosity >= self.NORMAL:
                    self.stdout.write(f" - {song}\n")
                self.imported_counter += 1
            else:
                if self.verbosity >= self.NORMAL:
                    self.stderr.write(
                        f"Errors importing song "
                        f"{row_dict['artist']} - {row_dict['title']}:\n"
                    )
                    self.stderr.write(f"{form.errors.as_json()}\n")
                self.skipped_counter += 1

    def finalize(self):
        if self.verbosity >= self.NORMAL:
            self.stdout.write(f"-------------------------\n")
            self.stdout.write(f"Songs imported: {self.imported_counter}\n")
            self.stdout.write(f"Songs skipped: {self.skipped_counter}\n\n")
