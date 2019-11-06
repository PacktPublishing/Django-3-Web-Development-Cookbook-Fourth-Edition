from copy import copy
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, builtins
from openpyxl.styles.numbers import FORMAT_NUMBER
from openpyxl.writer.excel import save_virtual_workbook

from django.contrib import admin
from django.db import models
from django.http.response import HttpResponse
from django.template.loader import render_to_string
from django.utils.html import mark_safe
from django.utils.translation import ugettext_lazy as _
from ordered_model.admin import OrderedTabularInline, OrderedInlineModelAdminMixin

from .models import Product, ProductPhoto


def export_xlsx(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    number_alignment = Alignment(horizontal="right")
    wb.add_named_style(
        NamedStyle(
            "Identifier", alignment=number_alignment, number_format=FORMAT_NUMBER
        )
    )
    wb.add_named_style(
        NamedStyle("Normal Wrapped", alignment=Alignment(wrap_text=True))
    )

    number_headline_1 = copy(builtins.styles["Headline 1"])
    number_headline_1.name = "Number Headline 1"
    number_headline_1.alignment = number_alignment
    wb.add_named_style(number_headline_1)

    class Config:
        def __init__(
            self,
            heading,
            width=None,
            heading_style="Headline 1",
            style="Normal Wrapped",
        ):
            self.heading = heading
            self.width = width
            self.heading_style = heading_style
            self.style = style

    column_config = {
        "A": Config(
            "ID", width=10, heading_style="Number Headline 1", style="Identifier"
        ),
        "B": Config("Title", width=30),
        "C": Config("Description", width=60),
        "D": Config(
            "Price ($)", width=15, heading_style="Number Headline 1", style="Currency"
        ),
        "E": Config("Preview", width=100, style="Hyperlink"),
    }

    # Set up column widths, header values and styles
    for col, conf in column_config.items():
        ws.column_dimensions[col].width = conf.width

        column = ws[f"{col}1"]
        column.value = conf.heading
        column.style = conf.heading_style

    # Add products
    for obj in queryset.order_by("pk"):
        project_photos = obj.productphoto_set.all()[:1]
        url = ""
        if project_photos:
            url = project_photos[0].photo.url

        data = [obj.pk, obj.title, obj.description, obj.price, url]
        ws.append(data)

        row = ws.max_row
        for row_cells in ws.iter_cols(min_row=row, max_row=row):
            for cell in row_cells:
                cell.style = column_config[cell.column].style

    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    charset = "utf-8"
    response = HttpResponse(
        content=save_virtual_workbook(wb),
        content_type=f"{mimetype}; charset={charset}",
        charset=charset,
    )
    response["Content-Disposition"] = "attachment; filename=products.xlsx"
    return response


export_xlsx.short_description = _("Export XLSX")


ZERO = "zero"
ONE = "one"
MANY = "many"


class PhotoFilter(admin.SimpleListFilter):
    # Human-readable title which will be displayed in the
    # right admin sidebar just above the filter options.
    title = _("photos")

    # Parameter for the filter that will be used in the
    # URL query.
    parameter_name = "photos"

    def lookups(self, request, model_admin):
        """
        Returns a list of tuples, akin to the values given for
        model field choices. The first element in each tuple is the
        coded value for the option that will appear in the URL
        query. The second element is the human-readable name for
        the option that will appear in the right sidebar.
        """
        return (
            (ZERO, _("Has no photos")),
            (ONE, _("Has one photo")),
            (MANY, _("Has more than one photo")),
        )

    def queryset(self, request, queryset):
        """
        Returns the filtered queryset based on the value
        provided in the query string and retrievable via
        `self.value()`.
        """
        qs = queryset.annotate(num_photos=models.Count("productphoto"))

        if self.value() == ZERO:
            qs = qs.filter(num_photos=0)
        elif self.value() == ONE:
            qs = qs.filter(num_photos=1)
        elif self.value() == MANY:
            qs = qs.filter(num_photos__gte=2)
        return qs


class ProductPhotoInline(OrderedTabularInline):
    model = ProductPhoto
    extra = 0
    fields = ("photo", "order", "move_up_down_links")
    readonly_fields = ("order", "move_up_down_links")
    ordering = ("order",)


class ProductAdmin(OrderedInlineModelAdminMixin, admin.ModelAdmin):
    list_display = ["first_photo", "title", "price"]
    list_display_links = ["first_photo", "title"]
    list_editable = ["price"]
    list_filter = ["price", PhotoFilter]

    actions = [export_xlsx]

    fieldsets = ((_("Product"), {"fields": ("title", "slug", "description", "price")}),)
    prepopulated_fields = {"slug": ("title",)}
    inlines = [ProductPhotoInline]

    def first_photo(self, obj):
        project_photos = obj.productphoto_set.all()[:1]
        if project_photos.count() > 0:
            photo_preview = render_to_string(
                "products/includes/photo-preview.html",
                {
                    "photo": project_photos[0],
                    "product": obj,
                    "target": "preview",
                },
            )
            return mark_safe(photo_preview)
        return ""

    first_photo.short_description = _("Preview")


admin.site.register(Product, ProductAdmin)
